import openpyxl
import json
import default
from termcolor import colored
import sys

def test():
    wb = openpyxl.load_workbook("data/data.xlsx", data_only=True)
    sheet = wb['result']

    c1 = sheet["B2"]
    c1.value = "Vishesh"

    print(c1.value)

    wb.save("data/data.xlsx")
    wb.close()

def TotalResultCountType(paper, type):
    wb = openpyxl.load_workbook("data/data.xlsx", data_only=True)
    sheet = wb['analysis']
    Number = 0

    for row in sheet.iter_rows(min_col=3, max_col=3):
        for cell in row:
            if str(cell.value) == paper:
                # print(cell, row, cell.row)
                if sheet.cell(row=cell.row, column=14).value == type:
                    Number+=1

    return Number

def TotalResultCountForm(paper, form):
    wb = openpyxl.load_workbook("data/data.xlsx", data_only=True)
    sheet = wb['analysis']
    Number = 0

    for row in sheet.iter_rows(min_col=3, max_col=3):
        for cell in row:
            if str(cell.value) == paper:
                if "Form"+str(sheet.cell(row=cell.row, column=16).value) == form:
                    Number+=1
    return Number

def weird_division(n, d):
    return n / d if d else 0

def printResult():
    wb = openpyxl.load_workbook("data/data.xlsx", data_only=True)
    sheet = wb['result']

    with open("data/json/Score.json", "r") as f:
        score = json.load(f)
        TotalScoreList = score["Score"]
    with open("data/json/TypeData.json", "r") as f:
        type = json.load(f)
    with open("data/json/FormData.json") as f:
        form = json.load(f)
    try:
        # Total Scores
        for row, (name, scores) in enumerate(TotalScoreList.items(), start=0):
            sheet.cell(row=row+2, column=2).value = scores

        # Name
        for row, (name, list) in enumerate(type.items(), start=0):
            sheet.cell(row=row+2, column=1).value = name
            # Type
            for col, (types, num) in enumerate(list.items(), start=0):
                if types != "year":
                    sheet.cell(row=row+2, column=col+8).value = num

        # Form
        for row, (name, list) in enumerate(form.items(), start=0):
            for col, (form, num) in enumerate(list.items(), start=0):
                if form != "year":
                    sheet.cell(row=row+2, column=col+2).value = num

        wb.save("data/data.xlsx")
        wb.close()
    except (PermissionError):
        print(colored("Error! Please save and close data.xlsx and try again!", "red"))
        sys.exit()

def printResultPercentage():
    wb = openpyxl.load_workbook("data/data.xlsx", data_only=True)
    sheet = wb['resultPercentage']

    with open("data/json/Score.json", "r") as f:
        score = json.load(f)
        TotalScoreList = score["Score"]
    with open("data/json/TypeData.json", "r") as f:
        type = json.load(f)
    with open("data/json/FormData.json") as f:
        form = json.load(f)
    try:
        # Total Scores
        for row, (name, scores) in enumerate(TotalScoreList.items(), start=0):
            sheet.cell(row=row+2, column=2).value = scores

        # Name
        for row, (name, list) in enumerate(type.items(), start=0):
            if name != "None_None(None)":
                sheet.cell(row=row+2, column=1).value = name
                years = str(type[name]["year"])
                # Type
                for col, (types, num) in enumerate(list.items(), start=0):
                    if types != "year":
                        x = (weird_division(int(num), TotalResultCountType(years, types)))*100
                        sheet.cell(row=row+2, column=col+8).value = x
                        print(years, name, types, ":", x)

        # Form
        for row, (name, list) in enumerate(form.items(), start=0):
            if name != "None_None(None)":
                years = str(form[name]["year"])
                for col, (forms, num) in enumerate(list.items(), start=0):
                    if forms != "year":
                        x = (weird_division(int(num), TotalResultCountForm(years, forms)))*100
                        sheet.cell(row=row+2, column=col+2).value = x
                        print(years, name, forms, ":", x)

        wb.save("data/data.xlsx")
        wb.close()
    except (PermissionError):
        print(colored("Error! Please save and close data.xlsx and try again!", "red"))
        sys.exit()
